"""
Import historical DailyReport data from quarterly .xlsx cash-book files.

File layout (openpyxl values_only=True, 0-indexed rows/columns):
  rows[0]  — header text ("Остаток за …")
  rows[1]  — opening balances: [1]=kaspi [2]=halyk [3]=cash [4]=usd
             (4 columns side-by-side, NOT the same layout as transaction rows)
  rows[2]  — account labels
  rows[3]  — empty
  rows[4]  — column headers (skip)
  rows[5…] — transaction rows (pos[0] is int) + summary + doctor section + balance section

Transaction columns:
  pos[1]  kaspi  income (>0) or expense (<0)
  pos[2]  kaspi  expense (<0)
  pos[3]  kaspi  comment
  pos[4]  halyk  income
  pos[5]  halyk  expense
  pos[6]  halyk  comment
  pos[7]  cash   income
  pos[8]  cash   expense
  pos[9]  cash   comment

Closing balance section detected by pos[5] in account-name set:
  pos[5]='Каспи pay'  → pos[6]=kaspi end-balance
  pos[5]='Halyk bank' → pos[6]=halyk end-balance
  pos[5]='Наличные'   → pos[6]=cash  end-balance
  pos[5]='USD'        → pos[6]=usd   end-balance
"""

import re
from datetime import datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path

from django.core.management.base import BaseCommand, CommandError
from django.db import transaction as db_transaction

try:
    import openpyxl
except ImportError:
    openpyxl = None

from apps.finance.models import DailyBalance, DailyReport, DailyTransaction

# ── column layout ──────────────────────────────────────────────────────────────
# Transaction rows (rows[5+]):  3 accounts × {income, expense, comment} columns
# (account_slug, income_col, expense_col, comment_col)
ACCOUNT_COLS = [
    ("kaspi_pay", 1, 2, 3),
    ("halyk",     4, 5, 6),
    ("cash",      7, 8, 9),
]

# Opening-balance row (rows[1]):  4 accounts laid out side-by-side
# (account_slug, col)
OPENING_ACCOUNT_COLS = [
    ("kaspi_pay", 1),
    ("halyk",     2),
    ("cash",      3),
    ("usd",       4),
]

# pos[5] value → account slug for end-balance section
BALANCE_LABEL = {
    "каспи pay":  "kaspi_pay",
    "halyk bank": "halyk",
    "наличные":   "cash",
    "usd":        "usd",
}

# pos[1] string values that mark the start of the doctor-payroll section
DOCTOR_MARKERS = {"врачи", "оплатили", "рентген", "проценты", "оплата"}


# ── helpers ────────────────────────────────────────────────────────────────────

def to_dec(val) -> Decimal | None:
    """Return Decimal if val is a non-zero number, otherwise None."""
    if val is None:
        return None
    if isinstance(val, str):
        return None
    try:
        d = Decimal(str(val))
        return d if d != 0 else None
    except (InvalidOperation, TypeError):
        return None


def _date_part(title: str) -> str | None:
    clean = title.strip()
    if not clean.startswith("Расчеты "):
        return None
    return clean[len("Расчеты "):].strip()


# Same-month range:  "30-31.05.26" / "7-8.08.25"
_RE_SAME_MONTH = re.compile(r"^(\d{1,2})-(\d{1,2})\.(\d{1,2})\.(\d{2})$")
# Cross-month range: "28.04-01.05.26"
_RE_CROSS_MONTH = re.compile(r"^(\d{1,2})\.(\d{1,2})-(\d{1,2})\.(\d{1,2})\.(\d{2})$")


def parse_sheet_date(title: str):
    """
    Return the LAST date of the sheet (use last day if range).

    Examples:
      "Расчеты 02.06.26"        → 2026-06-02
      "Расчеты 30-31.05.26"     → 2026-05-31
      "Расчеты 7-8.08.25"       → 2025-08-08
      "Расчеты 28.04-01.05.26"  → 2026-05-01  (cross-month)

    Returns None if unparseable.
    """
    date_part = _date_part(title)
    if date_part is None:
        return None

    # Single-day "dd.mm.yy"
    try:
        return datetime.strptime(date_part, "%d.%m.%y").date()
    except ValueError:
        pass

    # Cross-month "d.m-d.m.yy" → last day = (g3, g4, g5)
    m = _RE_CROSS_MONTH.match(date_part)
    if m:
        try:
            return datetime.strptime(f"{m.group(3)}.{m.group(4)}.{m.group(5)}", "%d.%m.%y").date()
        except ValueError:
            pass

    # Same-month "d-d.m.yy" → last day = (g2, g3, g4)
    m = _RE_SAME_MONTH.match(date_part)
    if m:
        try:
            return datetime.strptime(f"{m.group(2)}.{m.group(3)}.{m.group(4)}", "%d.%m.%y").date()
        except ValueError:
            pass

    return None


def get_first_date_of_range(title: str):
    """
    For range sheets, return the FIRST date (used to delete legacy imports).
    Returns None if the sheet is not a range.
    """
    date_part = _date_part(title)
    if date_part is None:
        return None

    m = _RE_CROSS_MONTH.match(date_part)
    if m:
        try:
            return datetime.strptime(f"{m.group(1)}.{m.group(2)}.{m.group(5)}", "%d.%m.%y").date()
        except ValueError:
            pass

    m = _RE_SAME_MONTH.match(date_part)
    if m:
        try:
            return datetime.strptime(f"{m.group(1)}.{m.group(3)}.{m.group(4)}", "%d.%m.%y").date()
        except ValueError:
            pass

    return None


def is_range_sheet(title: str) -> bool:
    return get_first_date_of_range(title) is not None


def parse_sheet(ws) -> dict:
    """
    Parse one worksheet.  Returns:
    {
        "date": date | None,
        "opening": {account: Decimal},
        "closing":  {account: Decimal},
        "transactions": [
            {account, direction, amount: Decimal, comment: str, row_order: int}
        ],
        "skipped_rows": int,
    }
    """
    rows = list(ws.iter_rows(values_only=True))
    result = {
        "date": parse_sheet_date(ws.title),
        "opening": {},
        "closing": {},
        "transactions": [],
        "skipped_rows": 0,
    }

    # ── opening balances (rows[1]) ─────────────────────────────────────────
    # Layout: [1]=kaspi  [2]=halyk  [3]=cash  [4]=usd
    if len(rows) > 1:
        ob = rows[1]
        for slug, col in OPENING_ACCOUNT_COLS:
            v = to_dec(ob[col] if len(ob) > col else None)
            if v is not None:
                result["opening"][slug] = v

    # ── transaction rows (rows[5..]) ───────────────────────────────────────
    row_order = 0
    for row in rows[5:]:
        if not row or len(row) < 4:
            continue

        row0 = row[0]

        # ── detect doctor-payroll section → stop collecting transactions ──
        if row0 is None and isinstance(row[1], str):
            if row[1].strip().lower() in DOCTOR_MARKERS:
                break

        # ── detect closing-balance section ────────────────────────────────
        if len(row) > 6 and isinstance(row[5], str):
            key = row[5].strip().lower()
            if key in BALANCE_LABEL:
                v = to_dec(row[6] if len(row) > 6 else None)
                if v is not None:
                    result["closing"][BALANCE_LABEL[key]] = v
                continue  # keep scanning for more balance rows

        # ── skip non-numbered rows (summary, labels, empty) ───────────────
        if not isinstance(row0, (int, float)):
            continue

        # ── parse transaction columns ─────────────────────────────────────
        for account, inc_col, exp_col, cmt_col in ACCOUNT_COLS:
            comment = ""
            if len(row) > cmt_col and isinstance(row[cmt_col], str):
                comment = row[cmt_col].strip()

            # income column
            inc = to_dec(row[inc_col] if len(row) > inc_col else None)
            if inc is not None:
                if inc > 0:
                    row_order += 1
                    result["transactions"].append({
                        "account":   account,
                        "direction": "income",
                        "amount":    inc,
                        "comment":   comment,
                        "row_order": row_order,
                    })
                else:
                    # negative value in income column → treat as expense
                    row_order += 1
                    result["transactions"].append({
                        "account":   account,
                        "direction": "expense",
                        "amount":    abs(inc),
                        "comment":   comment,
                        "row_order": row_order,
                    })

            # dedicated expense column
            exp = to_dec(row[exp_col] if len(row) > exp_col else None)
            if exp is not None:
                row_order += 1
                result["transactions"].append({
                    "account":   account,
                    "direction": "expense",
                    "amount":    abs(exp),
                    "comment":   comment,
                    "row_order": row_order,
                })

    return result


def save_sheet(parsed: dict, source_file: str, dry_run: bool) -> dict:
    """
    Persist one parsed sheet.  Returns stats dict.
    """
    d = parsed["date"]
    txns = parsed["transactions"]
    stats = {"date": d, "tx_saved": 0, "tx_skipped": 0, "skipped_reason": None}

    if d is None:
        stats["skipped_reason"] = "unparseable date"
        return stats

    if DailyReport.objects.filter(date=d).exists():
        stats["skipped_reason"] = "already imported"
        return stats

    if dry_run:
        stats["tx_saved"] = len(txns)
        return stats

    with db_transaction.atomic():
        report = DailyReport.objects.create(date=d, source_file=source_file)

        # transactions first (so we can compute totals)
        for t in txns:
            DailyTransaction.objects.create(
                report=report,
                account=t["account"],
                direction=t["direction"],
                amount=t["amount"],
                comment=t["comment"],
                row_order=t["row_order"],
                source="excel",
            )
            stats["tx_saved"] += 1

        # Compute per-account income / expense from saved transactions
        TRANS_ACCOUNTS = ("kaspi_pay", "halyk", "cash")
        totals = {a: {"income": Decimal("0"), "expense": Decimal("0")} for a in TRANS_ACCOUNTS}
        for t in txns:
            if t["account"] in totals:
                totals[t["account"]][t["direction"]] += t["amount"]

        # Write balances: balance_start (from Excel opening), balance_end = start + income - expense
        all_accounts = (
            set(parsed["opening"]) | set(parsed["closing"]) | set(TRANS_ACCOUNTS)
        )
        for account in all_accounts:
            start = parsed["opening"].get(account, Decimal("0"))
            if account in totals:
                end = start + totals[account]["income"] - totals[account]["expense"]
            else:
                # Account not in the transaction layout (e.g. USD) — use Excel closing if present
                end = parsed["closing"].get(account, start)

            DailyBalance.objects.update_or_create(
                report=report,
                account=account,
                defaults={"balance_start": start, "balance_end": end},
            )

    return stats


class Command(BaseCommand):
    help = "Импорт исторических данных кассовой книги из .xlsx файлов"

    def add_arguments(self, parser):
        group = parser.add_mutually_exclusive_group(required=True)
        group.add_argument(
            "--file",
            metavar="PATH",
            help="Путь к одному .xlsx файлу",
        )
        group.add_argument(
            "--dir",
            metavar="PATH",
            help="Путь к директории — импортируются все .xlsx файлы",
        )
        parser.add_argument(
            "--dry-run",
            action="store_true",
            help="Только парсить, не сохранять. Печатает статистику.",
        )
        parser.add_argument(
            "--reparse-ranges",
            action="store_true",
            help="Обработать только листы с диапазоном дат (например 30-31.05.26). "
                 "Удаляет ранее импортированные записи по ПЕРВОЙ дате диапазона.",
        )

    def handle(self, *args, **options):
        if openpyxl is None:
            raise CommandError("openpyxl не установлен. Выполни: pip install openpyxl")

        dry_run = options["dry_run"]
        reparse_ranges = options["reparse_ranges"]

        # collect files
        if options["file"]:
            files = [Path(options["file"])]
        else:
            d = Path(options["dir"])
            if not d.is_dir():
                raise CommandError(f"Директория не найдена: {d}")
            files = sorted(d.glob("*.xlsx"))

        if not files:
            self.stdout.write(self.style.WARNING("Файлы .xlsx не найдены."))
            return

        total_reports = 0
        total_tx = 0
        total_skipped = 0
        total_unparseable = 0

        for fpath in files:
            if not fpath.exists():
                self.stdout.write(self.style.ERROR(f"  Файл не найден: {fpath}"))
                continue

            self.stdout.write(f"\n{'─'*60}")
            self.stdout.write(f"📂  {fpath.name}")

            try:
                wb = openpyxl.load_workbook(fpath, data_only=True, read_only=True)
            except Exception as e:
                self.stdout.write(self.style.ERROR(f"  Ошибка открытия: {e}"))
                continue

            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]

                # In --reparse-ranges mode, only process range sheets and clean up legacy.
                if reparse_ranges:
                    first_date = get_first_date_of_range(sheet_name)
                    if first_date is None:
                        continue  # skip non-range sheets

                    if not dry_run:
                        # Remove legacy report keyed by the FIRST date of the range
                        # (only if it was imported from Excel — preserves manual entries)
                        deleted, _ = (
                            DailyReport.objects
                            .filter(date=first_date)
                            .exclude(source_file="")
                            .delete()
                        )
                        if deleted:
                            self.stdout.write(
                                f"  🗑  Удалён старый отчёт за {first_date} (из диапазона)"
                            )

                parsed = parse_sheet(ws)
                stats = save_sheet(parsed, fpath.name, dry_run)

                d = stats["date"]
                reason = stats["skipped_reason"]

                if reason == "already imported":
                    self.stdout.write(
                        f"  ⏭  {sheet_name.strip():<30}  {d}  — уже импортирован"
                    )
                    total_skipped += 1
                elif reason == "unparseable date":
                    self.stdout.write(
                        self.style.WARNING(f"  ⚠️  {sheet_name.strip():<30}  — не удалось распознать дату")
                    )
                    total_unparseable += 1
                else:
                    flag = "[dry-run] " if dry_run else ""
                    self.stdout.write(
                        self.style.SUCCESS(
                            f"  ✅  {sheet_name.strip():<30}  {d}  "
                            f"{flag}→ {stats['tx_saved']} транзакций"
                        )
                    )
                    total_reports += 1
                    total_tx += stats["tx_saved"]

            wb.close()

        self.stdout.write(f"\n{'═'*60}")
        if dry_run:
            self.stdout.write(self.style.WARNING("  *** DRY RUN — данные НЕ сохранены ***"))
        self.stdout.write(
            f"  Импортировано дней:    {total_reports}\n"
            f"  Транзакций сохранено:  {total_tx}\n"
            f"  Пропущено (дубли):     {total_skipped}\n"
            f"  Не распознана дата:    {total_unparseable}"
        )
