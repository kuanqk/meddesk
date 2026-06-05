"""
Перечитать opening balances из исходных xlsx и обновить
DailyBalance.balance_start для уже импортированных дней.

Используется для починки данных после бага в маппинге колонок opening row
(до фикса halyk и cash читались из неверных позиций).

НЕ трогает: balance_end, транзакции, прочие поля DailyReport/DailyBalance.
"""

from datetime import datetime
from decimal import Decimal
from pathlib import Path

from django.core.management.base import BaseCommand, CommandError

try:
    import openpyxl
except ImportError:
    openpyxl = None

from apps.finance.models import DailyBalance, DailyReport
from apps.finance.management.commands.import_excel import (
    OPENING_ACCOUNT_COLS,
    parse_sheet_date,
    to_dec,
)


def parse_opening_row(ws) -> dict[str, Decimal]:
    """Парсит только opening balances (rows[1]). Возвращает {account: Decimal}."""
    rows = list(ws.iter_rows(min_row=1, max_row=2, values_only=True))
    if len(rows) < 2:
        return {}
    ob = rows[1]
    result: dict[str, Decimal] = {}
    for slug, col in OPENING_ACCOUNT_COLS:
        v = to_dec(ob[col] if len(ob) > col else None)
        if v is not None:
            result[slug] = v
    return result


class Command(BaseCommand):
    help = "Обновить DailyBalance.balance_start из исходных xlsx (для починки бага маппинга)"

    def add_arguments(self, parser):
        parser.add_argument(
            "--dir",
            metavar="PATH",
            required=True,
            help="Директория с .xlsx файлами",
        )
        parser.add_argument(
            "--dry-run",
            action="store_true",
            help="Показать что изменится, не сохранять.",
        )

    def handle(self, *args, **options):
        if openpyxl is None:
            raise CommandError("openpyxl не установлен. Выполни: pip install openpyxl")

        dry_run = options["dry_run"]

        d = Path(options["dir"])
        if not d.is_dir():
            raise CommandError(f"Директория не найдена: {d}")
        files = sorted(d.glob("*.xlsx"))
        if not files:
            self.stdout.write(self.style.WARNING("Файлы .xlsx не найдены."))
            return

        if dry_run:
            self.stdout.write(self.style.WARNING("*** DRY RUN — изменения не сохраняются ***\n"))

        total_updates = 0
        total_unchanged = 0
        total_no_report = 0
        total_no_balance = 0
        total_unparseable = 0

        for fpath in files:
            self.stdout.write(f"\n📂 {fpath.name}")

            try:
                wb = openpyxl.load_workbook(fpath, data_only=True, read_only=True)
            except Exception as e:
                self.stdout.write(self.style.ERROR(f"  Ошибка открытия: {e}"))
                continue

            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                sheet_date = parse_sheet_date(sheet_name)

                if sheet_date is None:
                    total_unparseable += 1
                    continue

                opening = parse_opening_row(ws)
                if not opening:
                    continue

                # Найти отчёт за эту дату
                try:
                    report = DailyReport.objects.get(date=sheet_date)
                except DailyReport.DoesNotExist:
                    total_no_report += 1
                    continue

                # Обновить balance_start у существующих DailyBalance
                for account, new_start in opening.items():
                    # Пропустим USD на всякий случай — мы его не показываем в UI
                    if account == "usd":
                        continue

                    try:
                        bal = DailyBalance.objects.get(report=report, account=account)
                    except DailyBalance.DoesNotExist:
                        total_no_balance += 1
                        continue

                    if bal.balance_start == new_start:
                        total_unchanged += 1
                        continue

                    old = bal.balance_start
                    delta = new_start - old
                    self.stdout.write(
                        f"  {sheet_date}  {account:<10}  "
                        f"{old:>14,.2f} → {new_start:>14,.2f}  (Δ={delta:+,.2f})"
                    )

                    if not dry_run:
                        # Точечный update — не трогаем balance_end и др. поля
                        DailyBalance.objects.filter(pk=bal.pk).update(
                            balance_start=new_start,
                        )

                    total_updates += 1

            wb.close()

        self.stdout.write(f"\n{'═'*60}")
        verb = "будет обновлено" if dry_run else "обновлено"
        self.stdout.write(
            f"  Балансов {verb}:     {total_updates}\n"
            f"  Без изменений:         {total_unchanged}\n"
            f"  Нет DailyReport:       {total_no_report}\n"
            f"  Нет DailyBalance:      {total_no_balance}\n"
            f"  Не распознана дата:    {total_unparseable}"
        )
