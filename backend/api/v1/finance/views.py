import calendar
import logging
from datetime import date, datetime
from decimal import Decimal

from django.db import transaction as db_transaction
from django.db.models import Count, Max, Min, Q, Sum
from django.db.models.functions import TruncMonth
from django.http import HttpResponse
from django.utils import timezone
from rest_framework.response import Response
from rest_framework.views import APIView

from apps.accounts.permissions import (
    is_owner,
    is_owner_or_admin,
    require_tab,
)
from apps.finance.models import (
    DailyBalance,
    DailyReport,
    DailyTransaction,
    DoctorRevenue,
    PayrollCalculation,
)
from apps.finance.services.balances import opening_for_date, propagate_balances_forward
from apps.finance.services.sync import FinanceSyncService

from .serializers import (
    DailyBalanceSerializer,
    DailyReportSaveSerializer,
    DailySummarySerializer,
    ExpenseCategorySerializer,
    MonthlySummarySerializer,
    PayrollCalculationSerializer,
)

logger = logging.getLogger(__name__)

# ── helpers ────────────────────────────────────────────────────────────────────

def _parse_month(s: str) -> date:
    return datetime.strptime(s, "%Y-%m").date()


def _parse_date(s: str) -> date:
    return datetime.strptime(s, "%Y-%m-%d").date()


def _month_end(d: date) -> date:
    _, last = calendar.monthrange(d.year, d.month)
    return date(d.year, d.month, last)


def _next_month(d: date) -> date:
    if d.month == 12:
        return date(d.year + 1, 1, 1)
    return date(d.year, d.month + 1, 1)


def _iter_months(start: date, end: date):
    """Yield first-of-month dates from start..end inclusive."""
    cur = start.replace(day=1)
    ceil = end.replace(day=1)
    while cur <= ceil:
        yield cur
        cur = _next_month(cur)


def _all_data_bounds() -> tuple[date, date]:
    """Границы всей истории: (первое-число-месяца, конец-месяца) по DailyReport.

    Если данных нет — текущий месяц. Используется, когда from/to не заданы
    (режим «Всё»).
    """
    b = DailyReport.objects.aggregate(mn=Min("date"), mx=Max("date"))
    if b["mn"]:
        return b["mn"].replace(day=1), _month_end(b["mx"])
    today = date.today()
    return today.replace(day=1), _month_end(today)


# ── expense keyword categories ─────────────────────────────────────────────────

EXPENSE_CATEGORIES = [
    ("Зарплаты",   Q(comment__icontains="зп") | Q(comment__icontains="зарплат")),
    ("Материалы",  Q(comment__icontains="матер") | Q(comment__icontains="dental") | Q(comment__icontains="стом")),
    ("Аренда",     Q(comment__icontains="аренда") | Q(comment__icontains="мурзаб")),
    ("Комиссии",   Q(comment__icontains="комисс")),
    ("Наркоз",     Q(comment__icontains="наркоз") | Q(comment__icontains="калымбет")),
]

_ALL_KNOWN_Q = Q()
for _name, _q in EXPENSE_CATEGORIES:
    _ALL_KNOWN_Q |= _q


# ── views ──────────────────────────────────────────────────────────────────────

class FinanceSummaryView(APIView):
    """GET /api/v1/finance/summary/?from=YYYY-MM&to=YYYY-MM"""

    permission_classes = [require_tab("finance")]

    def get(self, request):
        qp = request.query_params
        default_from, default_to = _all_data_bounds()
        try:
            date_from = _parse_month(qp["from"]).replace(day=1) if qp.get("from") else default_from
            date_to = _month_end(_parse_month(qp["to"])) if qp.get("to") else default_to
        except ValueError:
            return Response({"error": "Use YYYY-MM format"}, status=400)

        # ── per-month income / expense sums ───────────────────────────────────
        def monthly_sum(direction: str) -> dict:
            qs = (
                DailyTransaction.objects
                .filter(direction=direction, report__date__range=[date_from, date_to])
                .annotate(month=TruncMonth("report__date"))
                .values("month")
                .annotate(total=Sum("amount"))
            )
            return {
                row["month"].date() if hasattr(row["month"], "date") else row["month"]: row["total"]
                for row in qs
            }

        income_map = monthly_sum("income")
        expense_map = monthly_sum("expense")

        # ── end-of-month closing balances ─────────────────────────────────────
        # For each month, grab balances from the last DailyReport of that month.
        results = []
        for month_start in _iter_months(date_from, date_to):
            month_end = _month_end(month_start)
            income = income_map.get(month_start, Decimal("0"))
            expenses = expense_map.get(month_start, Decimal("0"))

            kaspi = halyk = cash = None
            last_report = (
                DailyReport.objects
                .filter(date__range=[month_start, month_end])
                .prefetch_related("balances")
                .order_by("-date")
                .first()
            )
            if last_report:
                for bal in last_report.balances.all():
                    if bal.account == "kaspi_pay":
                        kaspi = bal.balance_end
                    elif bal.account == "halyk":
                        halyk = bal.balance_end
                    elif bal.account == "cash":
                        cash = bal.balance_end

            results.append({
                "month": month_start.strftime("%Y-%m"),
                "income": income,
                "expenses": expenses,
                "profit": income - expenses,
                "kaspi_balance_end": kaspi,
                "halyk_balance_end": halyk,
                "cash_balance_end": cash,
            })

        serializer = MonthlySummarySerializer(results, many=True)
        return Response(serializer.data)


class FinanceDailyView(APIView):
    """GET /api/v1/finance/daily/?from=YYYY-MM-DD&to=YYYY-MM-DD"""

    permission_classes = [require_tab("finance")]

    def get(self, request):
        try:
            date_from = _parse_date(request.query_params.get("from", ""))
            date_to = _parse_date(request.query_params.get("to", ""))
        except ValueError:
            return Response({"error": "Use YYYY-MM-DD format"}, status=400)

        # income / expenses per day
        def daily_sum(direction: str) -> dict:
            qs = (
                DailyTransaction.objects
                .filter(direction=direction, report__date__range=[date_from, date_to])
                .values("report__date")
                .annotate(total=Sum("amount"))
            )
            return {row["report__date"]: row["total"] for row in qs}

        income_map = daily_sum("income")
        expense_map = daily_sum("expense")

        # closing total balance per day from DailyBalance
        balance_qs = (
            DailyBalance.objects
            .filter(
                report__date__range=[date_from, date_to],
                account__in=["kaspi_pay", "halyk", "cash"],
            )
            .values("report__date")
            .annotate(total=Sum("balance_end"))
        )
        balance_map = {row["report__date"]: row["total"] for row in balance_qs}

        # collect all dates that have any data
        all_dates = sorted(
            set(income_map) | set(expense_map) | set(balance_map)
        )

        results = []
        for d in all_dates:
            inc = income_map.get(d, Decimal("0"))
            exp = expense_map.get(d, Decimal("0"))
            results.append({
                "date": d,
                "income": inc,
                "expenses": exp,
                "profit": inc - exp,
                "total_balance_end": balance_map.get(d),
            })

        serializer = DailySummarySerializer(results, many=True)
        return Response(serializer.data)


class FinanceExpensesView(APIView):
    """GET /api/v1/finance/expenses/?from=YYYY-MM&to=YYYY-MM"""

    permission_classes = [require_tab("finance")]

    def get(self, request):
        qp = request.query_params
        default_from, default_to = _all_data_bounds()
        try:
            date_from = _parse_month(qp["from"]).replace(day=1) if qp.get("from") else default_from
            date_to = _month_end(_parse_month(qp["to"])) if qp.get("to") else default_to
        except ValueError:
            return Response({"error": "Use YYYY-MM format"}, status=400)

        base_qs = DailyTransaction.objects.filter(
            direction="expense",
            report__date__range=[date_from, date_to],
        )
        grand_total = base_qs.aggregate(t=Sum("amount"))["t"] or Decimal("0")

        results = []
        for name, q_filter in EXPENSE_CATEGORIES:
            total = base_qs.filter(q_filter).aggregate(t=Sum("amount"))["t"] or Decimal("0")
            results.append((name, total))

        # Прочее = not matching any category
        other_total = base_qs.exclude(_ALL_KNOWN_Q).aggregate(t=Sum("amount"))["t"] or Decimal("0")
        results.append(("Прочее", other_total))

        data = [
            {
                "category": name,
                "total_amount": total,
                "percentage": float(total / grand_total * 100) if grand_total else 0.0,
            }
            for name, total in sorted(results, key=lambda x: x[1], reverse=True)
        ]

        serializer = ExpenseCategorySerializer(data, many=True)
        return Response(serializer.data)


class FinanceBalancesView(APIView):
    """GET /api/v1/finance/balances/?from=YYYY-MM-DD&to=YYYY-MM-DD"""

    permission_classes = [require_tab("finance")]

    def get(self, request):
        try:
            date_from = _parse_date(request.query_params.get("from", ""))
            date_to = _parse_date(request.query_params.get("to", ""))
        except ValueError:
            return Response({"error": "Use YYYY-MM-DD format"}, status=400)

        # Fetch all closing balances in range
        rows = (
            DailyBalance.objects
            .filter(
                report__date__range=[date_from, date_to],
                account__in=["kaspi_pay", "halyk", "cash"],
            )
            .values("report__date", "account", "balance_end")
            .order_by("report__date")
        )

        # Pivot: date → {account: balance}
        by_date: dict[date, dict] = {}
        for row in rows:
            d = row["report__date"]
            if d not in by_date:
                by_date[d] = {}
            by_date[d][row["account"]] = row["balance_end"]

        results = []
        for d in sorted(by_date):
            accs = by_date[d]
            kaspi = accs.get("kaspi_pay")
            halyk = accs.get("halyk")
            cash = accs.get("cash")
            total = sum(v for v in [kaspi, halyk, cash] if v is not None) or None
            results.append({
                "date": d,
                "kaspi": kaspi,
                "halyk": halyk,
                "cash": cash,
                "total": total,
            })

        serializer = DailyBalanceSerializer(results, many=True)
        return Response(serializer.data)


class DoctorsRevenueView(APIView):
    """GET /api/v1/finance/doctors-revenue/?from=YYYY-MM-DD&to=YYYY-MM-DD"""

    permission_classes = [require_tab("finance")]

    def get(self, request):
        try:
            date_from = _parse_date(request.query_params.get("from", ""))
            date_to = _parse_date(request.query_params.get("to", ""))
        except ValueError:
            return Response({"error": "Use YYYY-MM-DD format"}, status=400)

        rows = (
            DoctorRevenue.objects
            .filter(date__range=[date_from, date_to])
            .values("doctor", "doctor__name", "doctor__kpi_threshold")
            .annotate(
                revenue_total=Sum("revenue"),
                days_worked=Count("date", distinct=True),
            )
            .order_by("-revenue_total")
        )

        grand_total = sum((r["revenue_total"] or Decimal("0")) for r in rows) or Decimal("0")

        results = []
        for r in rows:
            total = r["revenue_total"] or Decimal("0")
            days = r["days_worked"] or 0
            per_day = (total / days) if days else Decimal("0")
            share = float(total / grand_total * 100) if grand_total else 0.0
            results.append({
                "doctor_id": r["doctor"],
                "doctor_name": r["doctor__name"],
                "revenue_total": f"{total:.2f}",
                "days_worked": days,
                "revenue_per_day": f"{per_day:.2f}",
                "share_percent": round(share, 1),
                "kpi_threshold": f"{(r['doctor__kpi_threshold'] or Decimal('0')):.2f}",
            })

        return Response(results)


# ── DailyReport CRUD ───────────────────────────────────────────────────────────

_BALANCE_ACCOUNTS = ["kaspi_pay", "halyk", "cash"]


def _build_report_response(report_date: date) -> dict:
    """Return the canonical GET payload for a given date."""
    ZERO = "0"

    try:
        report = (
            DailyReport.objects
            .prefetch_related("transactions", "balances")
            .get(date=report_date)
        )
        opening: dict[str, str] = {a: ZERO for a in _BALANCE_ACCOUNTS}
        for b in report.balances.all():
            if b.account in opening:
                opening[b.account] = str(b.balance_start)

        transactions = [
            {
                "id":        t.id,
                "account":   t.account,
                "direction": t.direction,
                "amount":    str(t.amount),
                "comment":   t.comment,
                "row_order": t.row_order,
                "source":    t.source,
            }
            for t in sorted(
                report.transactions.all(),
                key=lambda t: (t.account, t.row_order),
            )
        ]
        return {
            "date":             str(report_date),
            "exists":           True,
            "is_closed":        report.is_closed,
            "closed_by":        report.closed_by.username if report.closed_by else None,
            "closed_at":        report.closed_at.isoformat() if report.closed_at else None,
            "notes":            report.notes,
            "transactions":     transactions,
            "opening_balances": opening,
        }

    except DailyReport.DoesNotExist:
        # No report yet — opening is the most recent prior balance_end per
        # account (gap-day bridging), shared with the server-side chain.
        opening = {a: str(v) for a, v in opening_for_date(report_date).items()}

        return {
            "date":             str(report_date),
            "exists":           False,
            "is_closed":        False,
            "closed_by":        None,
            "closed_at":        None,
            "notes":            "",
            "transactions":     [],
            "opening_balances": opening,
        }


# Owner / owner-or-admin checks live in apps.accounts.permissions and are
# reused here (and by settings/views) — see is_owner / is_owner_or_admin.
_is_owner = is_owner
_is_owner_or_admin = is_owner_or_admin


class DailyReportView(APIView):
    """
    GET  /api/v1/finance/daily-report/?date=YYYY-MM-DD
    POST /api/v1/finance/daily-report/
    """

    permission_classes = [require_tab("finance")]

    def get(self, request):
        date_str = request.query_params.get("date")
        if not date_str:
            return Response({"error": "date parameter required"}, status=400)
        try:
            report_date = datetime.strptime(date_str, "%Y-%m-%d").date()
        except ValueError:
            return Response({"error": "Invalid date format. Use YYYY-MM-DD"}, status=400)

        return Response(_build_report_response(report_date))

    def post(self, request):
        serializer = DailyReportSaveSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        data = serializer.validated_data
        report_date = data["date"]

        with db_transaction.atomic():
            report, _ = DailyReport.objects.get_or_create(
                date=report_date,
                defaults={"created_by": request.user},
            )

            if report.is_closed:
                return Response(
                    {"error": "Отчёт закрыт и не может быть изменён."},
                    status=400,
                )

            report.notes = data.get("notes", "")
            report.save(update_fields=["notes", "updated_at"])

            # Replace only the manual rows. Imported rows (excel/macdent and
            # legacy null source) are left untouched so reopening + resaving an
            # imported day doesn't duplicate or drop them.
            report.transactions.filter(source="manual").delete()

            # Incoming rows are always stored as manual — the client must not
            # send back imported rows, and any source it provides is ignored.
            tx_objs = [
                DailyTransaction(
                    report=report,
                    account=tx["account"],
                    direction=tx["direction"],
                    amount=tx["amount"],
                    comment=tx.get("comment", ""),
                    row_order=i,
                    source="manual",
                )
                for i, tx in enumerate(data["transactions"])
            ]
            DailyTransaction.objects.bulk_create(tx_objs)

            # Server derives balance_start from the previous day's balance_end
            # (the cascade chain). The client's opening_balances are IGNORED for
            # computation — kept in the payload only for backward compat.
            # balance_end = balance_start + Σincome − Σexpense over ALL rows of
            # the report (manual + imported), as established in C3.
            opening = opening_for_date(report_date)
            client_opening = data.get("opening_balances") or {}
            all_tx = list(report.transactions.all())
            for account in _BALANCE_ACCOUNTS:
                has_prior = DailyBalance.objects.filter(
                    report__date__lt=report_date, account=account
                ).exists()
                if has_prior:
                    start = opening[account]
                else:
                    # Anchor (earliest day for this account): keep its stored
                    # balance_start; never zero it from a derived chain.
                    existing = DailyBalance.objects.filter(
                        report=report, account=account
                    ).first()
                    start = existing.balance_start if existing else Decimal("0")

                if account in client_opening and Decimal(str(client_opening[account])) != start:
                    logger.info(
                        "Ignoring client opening for %s on %s: client=%s server=%s",
                        account, report_date, client_opening[account], start,
                    )

                income = sum(
                    (t.amount for t in all_tx
                     if t.account == account and t.direction == "income"),
                    Decimal("0"),
                )
                expense = sum(
                    (t.amount for t in all_tx
                     if t.account == account and t.direction == "expense"),
                    Decimal("0"),
                )
                DailyBalance.objects.update_or_create(
                    report=report,
                    account=account,
                    defaults={"balance_start": start, "balance_end": start + income - expense},
                )

            # Cascade: recompute the balance chain for all later days so editing
            # a past day keeps the whole register self-consistent.
            propagate_balances_forward(from_date=report_date, commit=True)

        return Response(_build_report_response(report_date))


# ── Close / Reopen / Closed dates ─────────────────────────────────────────────

class DailyReportCloseView(APIView):
    """POST /api/v1/finance/daily-report/close/ — owner only."""

    permission_classes = [require_tab("finance")]

    def post(self, request):
        if not _is_owner(request.user):
            return Response({"error": "Только владелец может закрывать дни."}, status=403)

        date_str = request.data.get("date")
        if not date_str:
            return Response({"error": "date required"}, status=400)
        try:
            report_date = datetime.strptime(date_str, "%Y-%m-%d").date()
        except ValueError:
            return Response({"error": "Invalid date format. Use YYYY-MM-DD"}, status=400)

        try:
            report = DailyReport.objects.get(date=report_date)
        except DailyReport.DoesNotExist:
            return Response({"error": "Отчёт за этот день не найден."}, status=404)

        report.is_closed = True
        report.closed_by = request.user
        report.closed_at = timezone.now()
        report.save(update_fields=["is_closed", "closed_by", "closed_at", "updated_at"])

        return Response(_build_report_response(report_date))


class DailyReportReopenView(APIView):
    """POST /api/v1/finance/daily-report/reopen/ — owner only."""

    permission_classes = [require_tab("finance")]

    def post(self, request):
        if not _is_owner(request.user):
            return Response({"error": "Только владелец может переоткрывать дни."}, status=403)

        date_str = request.data.get("date")
        if not date_str:
            return Response({"error": "date required"}, status=400)
        try:
            report_date = datetime.strptime(date_str, "%Y-%m-%d").date()
        except ValueError:
            return Response({"error": "Invalid date format. Use YYYY-MM-DD"}, status=400)

        try:
            report = DailyReport.objects.get(date=report_date)
        except DailyReport.DoesNotExist:
            return Response({"error": "Отчёт за этот день не найден."}, status=404)

        report.is_closed = False
        report.closed_by = None
        report.closed_at = None
        report.save(update_fields=["is_closed", "closed_by", "closed_at", "updated_at"])

        return Response(_build_report_response(report_date))


class DailyReportClosedDatesView(APIView):
    """
    GET /api/v1/finance/daily-report/closed-dates/?month=YYYY-MM
    → list of closed-day date strings in that month.
    """

    permission_classes = [require_tab("finance")]

    def get(self, request):
        month_str = request.query_params.get("month")
        if not month_str:
            return Response({"error": "month parameter required"}, status=400)
        try:
            month_start = datetime.strptime(month_str, "%Y-%m").date().replace(day=1)
        except ValueError:
            return Response({"error": "Invalid month format. Use YYYY-MM"}, status=400)

        month_end = _month_end(month_start)

        dates = (
            DailyReport.objects
            .filter(date__range=[month_start, month_end], is_closed=True)
            .values_list("date", flat=True)
            .order_by("date")
        )
        return Response([d.isoformat() for d in dates])


# ── Payroll (ФОТ) ──────────────────────────────────────────────────────────────

class PayrollListView(APIView):
    """GET /api/v1/finance/payroll/?month=YYYY-MM — список расчётов ФОТ за месяц."""

    permission_classes = [require_tab("finance")]

    def get(self, request):
        month_str = request.query_params.get("month")
        if not month_str:
            return Response({"error": "month parameter required"}, status=400)
        try:
            period = _parse_month(month_str).replace(day=1)
        except ValueError:
            return Response({"error": "Use YYYY-MM format"}, status=400)

        qs = (
            PayrollCalculation.objects
            .filter(period=period)
            .select_related("staff_member", "confirmed_by")
            .order_by("staff_member__name")
        )
        return Response(PayrollCalculationSerializer(qs, many=True).data)


class PayrollCalculateView(APIView):
    """POST /api/v1/finance/payroll/calculate/ — owner/admin only.

    body: {"month": "YYYY-MM"}
    """

    permission_classes = [require_tab("finance")]

    def post(self, request):
        if not _is_owner_or_admin(request.user):
            return Response(
                {"error": "Только владелец или администратор может рассчитывать ФОТ."},
                status=403,
            )

        month_str = request.data.get("month")
        if not month_str:
            return Response({"error": "month required"}, status=400)
        try:
            period = _parse_month(month_str).replace(day=1)
        except ValueError:
            return Response({"error": "Use YYYY-MM format"}, status=400)

        FinanceSyncService().calculate_payroll(period.year, period.month)

        # Re-fetch with relations for a consistent, serialized response.
        qs = (
            PayrollCalculation.objects
            .filter(period=period)
            .select_related("staff_member", "confirmed_by")
            .order_by("staff_member__name")
        )
        return Response(PayrollCalculationSerializer(qs, many=True).data)


class PayrollConfirmView(APIView):
    """POST /api/v1/finance/payroll/{id}/confirm/ — owner only."""

    permission_classes = [require_tab("finance")]

    def post(self, request, pk):
        if not _is_owner(request.user):
            return Response(
                {"error": "Только владелец может подтверждать ФОТ."},
                status=403,
            )

        try:
            payroll = PayrollCalculation.objects.select_related(
                "staff_member", "confirmed_by"
            ).get(pk=pk)
        except PayrollCalculation.DoesNotExist:
            return Response({"error": "Расчёт ФОТ не найден."}, status=404)

        payroll.is_confirmed = True
        payroll.confirmed_by = request.user
        payroll.confirmed_at = timezone.now()
        payroll.save(update_fields=["is_confirmed", "confirmed_by", "confirmed_at"])

        return Response(PayrollCalculationSerializer(payroll).data)


class PayrollUnconfirmView(APIView):
    """POST /api/v1/finance/payroll/{id}/unconfirm/ — owner only.

    Снимает подтверждение, после чего расчёт снова можно пересчитать.
    """

    permission_classes = [require_tab("finance")]

    def post(self, request, pk):
        if not _is_owner(request.user):
            return Response(
                {"error": "Только владелец может снимать подтверждение ФОТ."},
                status=403,
            )

        try:
            payroll = PayrollCalculation.objects.select_related(
                "staff_member", "confirmed_by"
            ).get(pk=pk)
        except PayrollCalculation.DoesNotExist:
            return Response({"error": "Расчёт ФОТ не найден."}, status=404)

        payroll.is_confirmed = False
        payroll.confirmed_by = None
        payroll.confirmed_at = None
        payroll.save(update_fields=["is_confirmed", "confirmed_by", "confirmed_at"])

        return Response(PayrollCalculationSerializer(payroll).data)


class IncomeExportView(APIView):
    """GET /api/v1/finance/income/export/?from=YYYY-MM&to=YYYY-MM

    Отдаёт XLSX со сводкой дохода кассы (P&L) по месяцам: доход, расходы,
    прибыль + строка итога. from/to опциональны — без них берётся вся
    история (от первого до последнего дневного отчёта).
    """

    permission_classes = [require_tab("finance")]

    def get(self, request):
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

        bounds = DailyReport.objects.aggregate(mn=Min("date"), mx=Max("date"))
        if not bounds["mn"]:
            return Response({"error": "Нет данных для выгрузки."}, status=404)

        qp = request.query_params
        try:
            date_from = (
                _parse_month(qp["from"]).replace(day=1)
                if qp.get("from") else bounds["mn"].replace(day=1)
            )
            date_to = (
                _month_end(_parse_month(qp["to"]))
                if qp.get("to") else _month_end(bounds["mx"])
            )
        except ValueError:
            return Response({"error": "Use YYYY-MM format"}, status=400)

        def monthly_sum(direction: str) -> dict:
            qs = (
                DailyTransaction.objects
                .filter(direction=direction, report__date__range=[date_from, date_to])
                .annotate(month=TruncMonth("report__date"))
                .values("month")
                .annotate(total=Sum("amount"))
            )
            return {
                (row["month"].date() if hasattr(row["month"], "date") else row["month"]): row["total"]
                for row in qs
            }

        income_map = monthly_sum("income")
        expense_map = monthly_sum("expense")

        wb = Workbook()
        ws = wb.active
        ws.title = "Доходы по месяцам"

        headers = ["Месяц", "Доход", "Расходы", "Прибыль"]
        ws.append(headers)

        header_fill = PatternFill("solid", fgColor="2563EB")
        header_font = Font(bold=True, color="FFFFFF")
        thin = Side(style="thin", color="D1CEC6")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        money_fmt = "# ##0"

        for col, _ in enumerate(headers, start=1):
            c = ws.cell(row=1, column=col)
            c.fill = header_fill
            c.font = header_font
            c.alignment = Alignment(horizontal="center")
            c.border = border

        total_income = total_expense = Decimal("0")
        for month_start in _iter_months(date_from, date_to):
            income = income_map.get(month_start, Decimal("0")) or Decimal("0")
            expenses = expense_map.get(month_start, Decimal("0")) or Decimal("0")
            profit = income - expenses
            total_income += income
            total_expense += expenses

            ws.append([month_start.strftime("%Y-%m"), income, expenses, profit])
            r = ws.max_row
            for col in range(1, 5):
                cell = ws.cell(row=r, column=col)
                cell.border = border
                if col >= 2:
                    cell.number_format = money_fmt

        # ── строка итога ────────────────────────────────────────────────────
        ws.append(["ИТОГО", total_income, total_expense, total_income - total_expense])
        r = ws.max_row
        for col in range(1, 5):
            cell = ws.cell(row=r, column=col)
            cell.font = Font(bold=True)
            cell.border = border
            if col >= 2:
                cell.number_format = money_fmt

        widths = [12, 16, 16, 16]
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[chr(64 + i)].width = w
        ws.freeze_panes = "A2"

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = (
            f'attachment; filename="income_by_month_{date_from:%Y%m}_{date_to:%Y%m}.xlsx"'
        )
        wb.save(response)
        return response


class DoctorsRevenueExportView(APIView):
    """GET /api/v1/finance/doctors-revenue/export/?from=YYYY-MM-DD&to=YYYY-MM-DD

    XLSX с доходом врачей (DoctorRevenue) за указанный период: строка на
    врача (выручка, дни, выручка/день, доля) + строка итога.
    """

    permission_classes = [require_tab("finance")]

    def get(self, request):
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

        try:
            date_from = _parse_date(request.query_params.get("from", ""))
            date_to = _parse_date(request.query_params.get("to", ""))
        except ValueError:
            return Response({"error": "Use YYYY-MM-DD format"}, status=400)

        rows = (
            DoctorRevenue.objects
            .filter(date__range=[date_from, date_to])
            .values("doctor__name")
            .annotate(
                revenue_total=Sum("revenue"),
                days_worked=Count("date", distinct=True),
            )
            .order_by("-revenue_total")
        )
        grand_total = sum((r["revenue_total"] or Decimal("0")) for r in rows) or Decimal("0")

        wb = Workbook()
        ws = wb.active
        ws.title = "Доход врачей"

        ws.append([f"Доход врачей за период {date_from:%d.%m.%Y} — {date_to:%d.%m.%Y}"])
        ws.append([])

        headers = ["Врач", "Выручка", "Дней", "Выручка/день", "Доля, %"]
        header_row = ws.max_row + 1
        ws.append(headers)

        header_fill = PatternFill("solid", fgColor="2563EB")
        header_font = Font(bold=True, color="FFFFFF")
        thin = Side(style="thin", color="D1CEC6")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        money_fmt = "# ##0"

        ws.cell(row=1, column=1).font = Font(bold=True, size=13)
        for col in range(1, len(headers) + 1):
            c = ws.cell(row=header_row, column=col)
            c.fill = header_fill
            c.font = header_font
            c.alignment = Alignment(horizontal="center")
            c.border = border

        for r in rows:
            total = r["revenue_total"] or Decimal("0")
            days = r["days_worked"] or 0
            per_day = (total / days) if days else Decimal("0")
            share = float(total / grand_total * 100) if grand_total else 0.0
            ws.append([r["doctor__name"], total, days, per_day, round(share, 1)])
            rr = ws.max_row
            for col in range(1, len(headers) + 1):
                cell = ws.cell(row=rr, column=col)
                cell.border = border
                if col in (2, 4):
                    cell.number_format = money_fmt

        ws.append(["ИТОГО", grand_total, "", "", 100.0])
        rr = ws.max_row
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=rr, column=col)
            cell.font = Font(bold=True)
            cell.border = border
            if col in (2, 4):
                cell.number_format = money_fmt

        for i, w in enumerate([32, 16, 8, 16, 10], start=1):
            ws.column_dimensions[chr(64 + i)].width = w
        ws.freeze_panes = f"A{header_row + 1}"

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = (
            f'attachment; filename="doctors_revenue_{date_from:%Y%m%d}_{date_to:%Y%m%d}.xlsx"'
        )
        wb.save(response)
        return response


class ExcelImportView(APIView):
    """POST /api/v1/finance/import-excel/  (multipart: file=<xlsx>)

    Импорт дневных отчётов из квартального xlsx кассовой книги.
    Существующие дни не перезаписываются (пропускаются) — добавляются
    только новые. ?dry_run=1 — предпросмотр без записи. Owner-only.
    """

    permission_classes = [require_tab("finance")]

    def post(self, request):
        if not _is_owner(request.user):
            return Response(
                {"error": "Только владелец может импортировать данные."}, status=403
            )

        upload = request.FILES.get("file")
        if upload is None:
            return Response({"error": "Файл не передан (поле 'file')."}, status=400)

        dry_run = str(request.query_params.get("dry_run", "")).lower() in ("1", "true", "yes")

        import openpyxl
        from apps.finance.management.commands.import_excel import parse_sheet, save_sheet

        try:
            wb = openpyxl.load_workbook(upload, read_only=True, data_only=True)
        except Exception as e:
            return Response({"error": f"Не удалось открыть файл: {e}"}, status=400)

        imported: list[str] = []
        skipped: list[dict] = []
        tx_total = 0
        for ws in wb.worksheets:
            try:
                parsed = parse_sheet(ws)
            except Exception:
                continue
            if parsed.get("date") is None:
                continue  # не лист-день
            stats = save_sheet(parsed, source_file=upload.name, dry_run=dry_run)
            if stats.get("skipped_reason"):
                skipped.append({
                    "date": str(stats["date"]) if stats["date"] else ws.title,
                    "reason": stats["skipped_reason"],
                })
            else:
                imported.append(str(stats["date"]))
                tx_total += stats["tx_saved"]
        wb.close()

        imported_dates = sorted(imported)
        return Response({
            "dry_run": dry_run,
            "imported_days": len(imported_dates),
            "imported_dates": imported_dates,
            "date_range": [imported_dates[0], imported_dates[-1]] if imported_dates else None,
            "transactions_saved": tx_total,
            "skipped_days": len(skipped),
            "skipped": skipped[:100],
        })
